#!/usr/bin/env python3
"""
neis_writeback.py — 확정 세특(소스) → NEIS '과목세부능력및특기사항' 즉시업로드 xlsx 이관.

Step 6 검수 테이블에서 확정된 세특을, NEIS에서 내려받은 과목세특 서식 xlsx의
J열(세부능력 및 특기사항)에 **결정론적으로** 써넣는다. 이 xlsx는 그대로 NEIS에
재업로드 가능한 공식 서식이다.

설계 원칙 (헌법 1 = 오매핑 0 을 write-back 경계까지 연장):
  - 이중키 대조: (반, 번호) 복합키 + 성명 **둘 다** 일치해야 기입.
  - non-empty 충돌 가드: 타깃 J셀에 소스와 '다른' 내용이 이미 있으면 기본은
    전체 중단(HALT) + 표면화. 맹목적 덮어쓰기 금지(헌법 6 = 교사 검수 관문 보호).
  - 백업 우선: 쓰기 전 원본 전량 백업.
  - 라운드트립 검증: 쓴 뒤 모든 셀을 다시 읽어 소스와 글자 대조.
  - 양방향 대사(reconcile): 소스에만/타깃에만 있는 학생을 리포트.

기본은 DRY-RUN(리포트만). 실제 기입은 --write.

NEIS 과목세특 서식 스키마 (2026 고등, 자동감지 — 열 위치 아닌 헤더명으로 매칭):
  시트 1개, 1행=헤더, 2행부터 데이터.
  학년도 | 학기 | 학년 | 학생개인번호 | 과목 | 과목코드 |
  반/번호("4/1") | 성명 | 학적변동구분 | **세부능력 및 특기사항(=기입 타깃)** | 영재·발명교육
  반별 1파일 (…_과목세특_4.xlsx, _5.xlsx, …).

사용 예:
  # 1) 대사만 (아무것도 안 씀)
  python3 neis_writeback.py --source out/명렬표-한국사-2026-세특.csv \
      --targets "~/Downloads/2026_1학기_1학년_1_한국사1_과목세특_*.xlsx"

  # 2) 실제 기입 (충돌 있으면 자동 중단)
  python3 neis_writeback.py --source ... --targets "..." --write

  # 3) 충돌 셀은 기존 유지하고 나머지만 기입
  python3 neis_writeback.py --source ... --targets "..." --write --on-conflict skip
"""
import argparse, csv, glob, json, re, shutil, sys, warnings
from datetime import datetime
from pathlib import Path

warnings.filterwarnings("ignore")  # openpyxl default-style warning
try:
    import openpyxl
    from openpyxl.styles import Alignment
except ImportError:
    sys.exit("openpyxl 필요: pip install openpyxl")

RESET = "\033[0m"; RED = "\033[31m"; GRN = "\033[32m"; YEL = "\033[33m"; CYN = "\033[36m"


def norm(s):
    """헤더/이름 정규화: None→'', 공백·개행 접기, strip."""
    return re.sub(r"\s+", "", str(s or ""))


def parse_key_from_hakbeon(hb):
    """5자리 학번 10401 → (반=4, 번호=1). 실패 시 None."""
    hb = re.sub(r"\D", "", str(hb or ""))
    if len(hb) == 5:
        return int(hb[1:3]), int(hb[3:5])
    return None


def parse_key_from_bannum(s):
    """'4/1' 또는 '4-1' → (4, 1). 실패 시 None."""
    m = re.match(r"\s*(\d+)\s*[/\-]\s*(\d+)\s*$", str(s or ""))
    return (int(m.group(1)), int(m.group(2))) if m else None


# ---------- 소스 로드 ----------
def load_source(path, id_col, name_col, content_col, class_col, num_col,
                status_col, status_ok):
    """소스(csv/json) → {(반,번호): {'name','content','raw'}}. 중복키는 에러."""
    p = Path(path)
    rows = []
    if p.suffix.lower() == ".json":
        data = json.loads(p.read_text(encoding="utf-8"))
        rows = data if isinstance(data, list) else data.get("students", data.get("rows", []))
    else:
        with open(p, encoding="utf-8-sig", newline="") as f:
            rows = list(csv.DictReader(f))
    if not rows:
        sys.exit(f"소스가 비었음: {path}")

    out = {}
    dupes = []
    skipped_status = 0
    for r in rows:
        if status_col and status_ok:
            if status_ok not in str(r.get(status_col, "")):
                skipped_status += 1
                continue
        # 키 결정: (class_col,num_col) 우선, 없으면 id_col(학번) 파싱
        key = None
        if class_col and num_col and r.get(class_col) not in (None, "") and r.get(num_col) not in (None, ""):
            try:
                key = (int(re.sub(r"\D", "", str(r[class_col]))), int(re.sub(r"\D", "", str(r[num_col]))))
            except ValueError:
                key = None
        if key is None and id_col:
            key = parse_key_from_hakbeon(r.get(id_col))
        if key is None:
            sys.exit(f"소스 행에서 반/번호 키를 못 구함: {r}")
        name = str(r.get(name_col, "")).strip()
        content = str(r.get(content_col, "")).strip()
        if key in out:
            dupes.append(key)
        out[key] = {"name": name, "content": content, "raw": r}
    if dupes:
        sys.exit(f"{RED}소스에 중복 (반,번호) 키: {dupes}{RESET}")
    return out, skipped_status


# ---------- 타깃 스키마 자동감지 ----------
def detect_columns(ws):
    """헤더행(1행)에서 반/번호·성명·세특 열 인덱스를 헤더명으로 찾음."""
    hdr = {c: norm(ws.cell(row=1, column=c).value) for c in range(1, ws.max_column + 1)}
    key_c = name_c = content_c = None
    for c, h in hdr.items():
        if h in ("반/번호", "반번호"):
            key_c = c
        elif h in ("성명", "이름"):
            name_c = c
        elif "세부능력" in h:  # '세부능력및특기사항'
            content_c = c
    return key_c, name_c, content_c


# ---------- 메인 ----------
def main():
    ap = argparse.ArgumentParser(description="확정 세특 → NEIS 과목세특 xlsx 결정론 이관")
    ap.add_argument("--source", required=True, help="확정 세특 소스 (csv 또는 json)")
    ap.add_argument("--targets", required=True, help="NEIS xlsx glob (따옴표로 감쌀 것)")
    ap.add_argument("--write", action="store_true", help="실제 기입 (미지정=dry-run)")
    ap.add_argument("--on-conflict", choices=["halt", "skip", "overwrite"], default="halt",
                    help="타깃 셀에 소스와 다른 기존 내용이 있을 때: "
                         "halt=전체중단(기본) / skip=기존유지 / overwrite=덮어쓰기")
    ap.add_argument("--no-backup", action="store_true", help="쓰기 전 백업 생략(비권장)")
    ap.add_argument("--no-wrap", action="store_true", help="자동 줄바꿈 서식 미적용")
    # 소스 컬럼명 (기본=본 프로젝트 명렬표 스키마)
    ap.add_argument("--id-col", default="학번", help="학번(5자리) 컬럼명")
    ap.add_argument("--name-col", default="이름")
    ap.add_argument("--content-col", default="세특")
    ap.add_argument("--class-col", default=None, help="반 컬럼명(학번 대신 반+번호로 키잡을 때)")
    ap.add_argument("--num-col", default=None, help="번호 컬럼명")
    ap.add_argument("--status-col", default=None, help="상태 필터 컬럼명(선택)")
    ap.add_argument("--status-ok", default=None, help="이 문자열 포함 행만 기입(예: 확정)")
    args = ap.parse_args()

    src, skipped_status = load_source(args.source, args.id_col, args.name_col,
                                      args.content_col, args.class_col, args.num_col,
                                      args.status_col, args.status_ok)
    files = sorted(glob.glob(args.targets))
    if not files:
        sys.exit(f"타깃 없음: {args.targets}")

    print(f"{CYN}소스 학생 {len(src)}명 | 타깃 파일 {len(files)}개"
          + (f" | 상태필터로 제외 {skipped_status}행" if skipped_status else "") + RESET)

    plan = []          # (file, row, key, name, action, csv_content)
    conflicts = []     # (file, key, name, existing, new)
    name_mismatch = [] # (file, key, xlsx_name, src_name)
    tgt_not_in_src = []# (file, key, name)
    seen_keys = set()

    for fn in files:
        wb = openpyxl.load_workbook(fn)
        ws = wb.active
        kc, nc, cc = detect_columns(ws)
        if not all([kc, nc, cc]):
            sys.exit(f"{RED}{Path(fn).name}: NEIS 스키마 감지 실패 "
                     f"(반/번호={kc} 성명={nc} 세특={cc}){RESET}")
        for r in range(2, ws.max_row + 1):
            g = ws.cell(row=r, column=kc).value
            name = ws.cell(row=r, column=nc).value
            if g in (None, "") and name in (None, ""):
                continue
            key = parse_key_from_bannum(g)
            if key is None:
                sys.exit(f"{RED}{Path(fn).name} row{r}: 반/번호 파싱 실패 {g!r}{RESET}")
            seen_keys.add(key)
            if key not in src:
                tgt_not_in_src.append((fn, key, name))
                continue
            s = src[key]
            if norm(s["name"]) != norm(name):
                name_mismatch.append((fn, key, name, s["name"]))
                continue  # 이름 불일치 = 오매핑 위험 → 절대 안 씀
            existing = ws.cell(row=r, column=cc).value
            new = s["content"]
            if existing not in (None, "") and str(existing).strip() != new.strip():
                conflicts.append((fn, key, name, str(existing).strip(), new))
            plan.append((fn, r, key, name, "write", new, kc, nc, cc))

    src_not_in_tgt = [(k, v["name"]) for k, v in src.items() if k not in seen_keys]

    # ---- 리포트 ----
    def loc(k): return f"{k[0]}/{k[1]}"
    print(f"\n{CYN}=== 대사(reconcile) ==={RESET}")
    print(f"  이중키 매칭 예정: {GRN}{len(plan)}{RESET}")
    print(f"  이름 불일치(미기입): {YEL}{len(name_mismatch)}{RESET}")
    for fn, k, xn, sn in name_mismatch:
        print(f"    ❌ {Path(fn).name} {loc(k)}: xlsx='{xn}' vs 소스='{sn}'")
    print(f"  타깃에만 있고 소스에 없음: {YEL}{len(tgt_not_in_src)}{RESET}")
    for fn, k, n in tgt_not_in_src:
        print(f"    ⚠️ {Path(fn).name} {loc(k)} {n}")
    print(f"  소스에만 있고 타깃에 없음: {YEL}{len(src_not_in_tgt)}{RESET}")
    for k, n in src_not_in_tgt:
        print(f"    ⚠️ {loc(k)} {n}")
    print(f"  기존내용 충돌: {YEL}{len(conflicts)}{RESET}")
    for fn, k, n, ex, nw in conflicts:
        print(f"    🔶 {Path(fn).name} {loc(k)} {n}")
        print(f"        기존: {ex[:70]}…")
        print(f"        소스: {nw[:70]}…")

    # ---- 쓰기 가부 판단 ----
    if not args.write:
        print(f"\n{CYN}DRY-RUN — 실제 기입하려면 --write{RESET}")
        return
    if conflicts and args.on_conflict == "halt":
        sys.exit(f"\n{RED}충돌 {len(conflicts)}건 → 중단(halt). "
                 f"검토 후 --on-conflict skip(기존유지) 또는 overwrite(덮어쓰기) 지정.{RESET}")

    conflict_keys = {(fn, k) for fn, k, *_ in conflicts}

    # ---- 백업 ----
    if not args.no_backup:
        stamp = datetime.now().strftime("%Y%m%d-%H%M%S")
        bdir = Path(files[0]).parent / f"_backup_neis_writeback_{stamp}"
        bdir.mkdir(exist_ok=True)
        for fn in files:
            shutil.copy2(fn, bdir / Path(fn).name)
        print(f"\n{GRN}백업: {bdir}{RESET}")

    # ---- 기입 ----
    wrap = Alignment(wrap_text=True, vertical="top")
    by_file = {}
    for item in plan:
        by_file.setdefault(item[0], []).append(item)
    written = skipped = 0
    for fn, items in by_file.items():
        wb = openpyxl.load_workbook(fn)
        ws = wb.active
        for fn_, r, key, name, _, new, kc, nc, cc in items:
            if (fn_, key) in conflict_keys and args.on_conflict == "skip":
                skipped += 1
                continue
            cell = ws.cell(row=r, column=cc)
            cell.value = new
            if not args.no_wrap:
                cell.alignment = wrap
            written += 1
        wb.save(fn)
    print(f"{GRN}기입 {written} | 충돌셀 기존유지 {skipped}{RESET}")

    # ---- 라운드트립 검증 ----
    ok = bad = 0
    for fn, items in by_file.items():
        wb = openpyxl.load_workbook(fn)
        ws = wb.active
        for fn_, r, key, name, _, new, kc, nc, cc in items:
            if (fn_, key) in conflict_keys and args.on_conflict == "skip":
                continue
            got = ws.cell(row=r, column=cc).value
            if str(got or "").strip() == new.strip():
                ok += 1
            else:
                bad += 1
                print(f"    {RED}❌ 라운드트립 불일치 {Path(fn).name} {key[0]}/{key[1]} {name}{RESET}")
    tag = GRN if bad == 0 else RED
    print(f"{tag}라운드트립 검증: OK {ok} | 불일치 {bad}{RESET}")
    if bad:
        sys.exit(1)


if __name__ == "__main__":
    main()
